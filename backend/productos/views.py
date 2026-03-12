from rest_framework import viewsets, filters
from django_filters.rest_framework import DjangoFilterBackend
from .models import Producto
from .serializers import ProductoSerializer
from rest_framework.permissions import IsAuthenticated
from .permissions import IsAdminOrReadOnly
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from rest_framework.decorators import api_view, permission_classes
from catalogos.models import Marca, Modelo, Color, Talla
from reportlab.pdfgen import canvas
from django.shortcuts import get_object_or_404

class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'marca': ['exact'],
        'modelo': ['exact'],
        'color': ['exact'],
        'talla': ['exact'],
        'marca__nombre': ['exact'],
        'modelo__nombre': ['exact'],
        'color__nombre': ['exact'],
        'talla__nombre': ['exact'],
    }
    search_fields = [
    'nombre',
    'marca__nombre',
    'modelo__nombre',
    'color__nombre',
    'talla__nombre',
    ]   
    ordering_fields = ['precio', 'nombre']

def exportar_productos_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    ws.append([
        "ID",
        "Nombre",
        "Marca",
        "Modelo",
        "Color",
        "Talla",
        "Precio",
    ])

    productos = Producto.objects.all()

    for producto in productos:
        ws.append([
            producto.id,
            producto.nombre,
            producto.marca.nombre,
            producto.modelo.nombre,
            producto.color.nombre,
            producto.talla.nombre,
            float(producto.precio),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="productos.xlsx"'

    wb.save(response)
    return response

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def cargar_productos_excel(request):
    archivo = request.FILES.get('archivo')

    if not archivo:
        return HttpResponse("No se envió ningún archivo", status=400)

    wb = load_workbook(archivo)
    ws = wb.active

    filas_creadas = 0

    for fila in ws.iter_rows(min_row=2, values_only=True):
        nombre, marca_nombre, modelo_nombre, color_nombre, talla_nombre, precio = fila

        if not nombre:
            continue

        marca, _ = Marca.objects.get_or_create(nombre=str(marca_nombre).strip())
        modelo, _ = Modelo.objects.get_or_create(nombre=str(modelo_nombre).strip())
        color, _ = Color.objects.get_or_create(nombre=str(color_nombre).strip())
        talla, _ = Talla.objects.get_or_create(nombre=str(talla_nombre).strip())

        Producto.objects.create(
            nombre=str(nombre).strip(),
            marca=marca,
            modelo=modelo,
            color=color,
            talla=talla,
            precio=precio,
        )

        filas_creadas += 1

    return HttpResponse(f"Se importaron {filas_creadas} productos correctamente ✅")

def ficha_producto_pdf(request, pk):
    producto = get_object_or_404(Producto, pk=pk)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="producto_{producto.id}.pdf"'

    p = canvas.Canvas(response)
    p.setTitle(f"Ficha Producto {producto.nombre}")

    y = 800

    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, y, "Ficha de Producto")

    y -= 40
    p.setFont("Helvetica", 12)
    p.drawString(100, y, f"ID: {producto.id}")

    y -= 25
    p.drawString(100, y, f"Nombre: {producto.nombre}")

    y -= 25
    p.drawString(100, y, f"Marca: {producto.marca.nombre}")

    y -= 25
    p.drawString(100, y, f"Modelo: {producto.modelo.nombre}")

    y -= 25
    p.drawString(100, y, f"Color: {producto.color.nombre}")

    y -= 25
    p.drawString(100, y, f"Talla: {producto.talla.nombre}")

    y -= 25
    p.drawString(100, y, f"Precio: S/ {producto.precio}")

    if producto.imagen:
        y -= 40
        p.drawString(100, y, f"Imagen: {producto.imagen.url}")

    p.showPage()
    p.save()

    return response